
from openpyxl import Workbook
import datetime as dt

wb = Workbook()

ws = wb.active

ws['A1'] = 42

ws.append([1,2,3])

ws['A3'] = dt.datetime.now()
ws.column_dimensions['A'].width = 20


wb.save('sample.xlsx')
